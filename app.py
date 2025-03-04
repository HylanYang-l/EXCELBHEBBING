from flask import Flask, render_template, request, send_file
import openpyxl
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

def column_name_to_number(name):
    result = 0
    for char in name.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def is_valid_column_name(name):
    if not name or not name.strip():
        return False
    return all(c.upper() in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ' for c in name)

def process_excel(file_path, reference_col, process_cols):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 将列名转换为列号
    try:
        ref_col = column_name_to_number(reference_col)
        process_col_nums = [column_name_to_number(col.strip()) for col in process_cols]
    except:
        raise ValueError("无效的列名")
    
    # 获取参考列的合并单元格
    merged_ranges = [r for r in sheet.merged_cells.ranges 
                    if r.min_col == ref_col and r.max_col == ref_col]
    
    for merged_range in merged_ranges:
        start_row = merged_range.min_row
        end_row = merged_range.max_row
        
        # 处理选中的列
        for col in process_col_nums:
            if col != ref_col:  # 跳过参考列
                # 收集对应列的所有非空值
                values = []
                for row in range(start_row, end_row + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value:
                        values.append(str(cell_value))
                
                # 合并内容并添加换行
                combined_value = '\n'.join(values)
                
                # 执行合并单元格操作
                col_letter = get_column_letter(col)
                merge_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                sheet.merge_cells(merge_range)
                
                # 设置合并后的值并配置单元格格式
                cell = sheet.cell(row=start_row, column=col)
                cell.value = combined_value
                alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
                cell.alignment = alignment
    
    output_path = file_path.replace('.xlsx', '_processed.xlsx')
    wb.save(output_path)
    return output_path

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return '没有选择文件'
        file = request.files['file']
        if file.filename == '':
            return '没有选择文件'
        
        # 获取并验证用户输入的列名
        reference_col = request.form.get('reference_col', 'A').strip()
        process_cols = [col.strip() for col in request.form.get('process_cols', '').split(',')]
        
        # 验证列名
        if not is_valid_column_name(reference_col):
            return '无效的参考列名'
        if not all(is_valid_column_name(col) for col in process_cols):
            return '无效的处理列名'
        if not process_cols:
            return '请至少选择一个要处理的列'
            
        if file and file.filename.endswith('.xlsx'):
            try:
                file_path = os.path.join(app.root_path, 'uploads', file.filename)
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                file.save(file_path)
                processed_file = process_excel(file_path, reference_col, process_cols)
                return send_file(processed_file, as_attachment=True)
            except ValueError as e:
                return str(e)
            except Exception as e:
                return f'处理出错：{str(e)}'
    return render_template('upload.html')

if __name__ == '__main__':
    os.makedirs(os.path.join(app.root_path, 'uploads'), exist_ok=True)
    app.run(debug=True)