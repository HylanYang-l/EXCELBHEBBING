import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.utils import get_column_letter
import os

class ExcelProcessor:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Excel表格处理工具")
        self.window.geometry("600x400")
        self.window.resizable(False, False)
        
        # 设置样式
        style = ttk.Style()
        style.configure('TButton', padding=6)
        style.configure('TLabel', padding=5)
        style.configure('TEntry', padding=5)
        
        self.create_widgets()
        
    def create_widgets(self):
        # 文件选择框
        file_frame = ttk.LabelFrame(self.window, text="选择文件", padding=10)
        file_frame.pack(fill="x", padx=20, pady=10)
        
        self.file_path = tk.StringVar()
        self.file_entry = ttk.Entry(file_frame, textvariable=self.file_path, width=50)
        self.file_entry.pack(side="left", padx=5)
        
        browse_btn = ttk.Button(file_frame, text="浏览", command=self.browse_file)
        browse_btn.pack(side="left", padx=5)
        
        # 列选择框
        cols_frame = ttk.LabelFrame(self.window, text="列设置", padding=10)
        cols_frame.pack(fill="x", padx=20, pady=10)
        
        # 参考列
        ref_label = ttk.Label(cols_frame, text="参考列:")
        ref_label.pack(anchor="w")
        self.ref_col = ttk.Entry(cols_frame, width=10)
        self.ref_col.insert(0, "A")
        self.ref_col.pack(anchor="w", padx=5, pady=2)
        
        # 处理列
        process_label = ttk.Label(cols_frame, text="处理列(用逗号分隔):")
        process_label.pack(anchor="w")
        self.process_cols = ttk.Entry(cols_frame, width=30)
        self.process_cols.insert(0, "B,C")
        self.process_cols.pack(anchor="w", padx=5, pady=2)
        
        # 处理按钮
        process_btn = ttk.Button(self.window, text="开始处理", command=self.process_file)
        process_btn.pack(pady=20)
        
        # 状态标签
        self.status_label = ttk.Label(self.window, text="")
        self.status_label.pack(pady=10)
        
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if filename:
            self.file_path.set(filename)
            
    def process_file(self):
        if not self.file_path.get():
            messagebox.showerror("错误", "请选择Excel文件")
            return
            
        try:
            self.status_label.config(text="处理中...")
            self.window.update()
            
            wb = openpyxl.load_workbook(self.file_path.get())
            sheet = wb.active
            
            ref_col = self.ref_col.get().strip().upper()
            process_cols = [col.strip().upper() for col in self.process_cols.get().split(',')]
            
            # 转换列名为列号
            ref_col_num = ord(ref_col) - ord('A') + 1
            process_col_nums = [ord(col) - ord('A') + 1 for col in process_cols]
            
            # 获取参考列的合并单元格
            merged_ranges = [r for r in sheet.merged_cells.ranges 
                           if r.min_col == ref_col_num and r.max_col == ref_col_num]
            
            for merged_range in merged_ranges:
                start_row = merged_range.min_row
                end_row = merged_range.max_row
                
                for col in process_col_nums:
                    values = []
                    for row in range(start_row, end_row + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value:
                            values.append(str(cell_value))
                    
                    combined_value = '\n'.join(values)
                    col_letter = get_column_letter(col)
                    merge_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                    sheet.merge_cells(merge_range)
                    
                    cell = sheet.cell(row=start_row, column=col)
                    cell.value = combined_value
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
            
            # 保存文件
            output_path = self.file_path.get().replace('.xlsx', '_processed.xlsx')
            wb.save(output_path)
            
            self.status_label.config(text=f"处理完成！文件已保存为: {output_path}")
            messagebox.showinfo("成功", "文件处理完成！")
            
        except Exception as e:
            self.status_label.config(text="处理出错")
            messagebox.showerror("错误", str(e))
    
    def run(self):
        self.window.mainloop()

if __name__ == "__main__":
    app = ExcelProcessor()
    app.run()