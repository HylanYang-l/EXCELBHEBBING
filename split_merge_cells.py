import openpyxl
from openpyxl.utils import get_column_letter

def process_merged_cells(filename):
    # 加载工作簿
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    
    # 获取所有合并单元格的范围
    merged_ranges = list(sheet.merged_cells.ranges)
    
    # 先解除所有合并的单元格
    for merged_range in merged_ranges:
        # 保存合并单元格中的值
        value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        # 解除合并
        sheet.unmerge_cells(str(merged_range))
        # 将值填充到解除合并的所有单元格中
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                sheet.cell(row=row, column=col, value=value)
    
    # 保存修改后的文件
    wb.save(filename.replace('.xlsx', '_processed.xlsx'))
    print(f"处理完成，文件已保存为: {filename.replace('.xlsx', '_processed.xlsx')}")

# 使用示例
if __name__ == "__main__":
    excel_file = "c:\\Users\\83406\\Desktop\\EXCELBHEBBING\\your_excel_file.xlsx"
    process_merged_cells(excel_file)